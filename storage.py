"""
Data storage module -- Excel-based data pipeline.
The ONLY module that reads or writes Excel files.
Replaces the old database.py (SQLite) with openpyxl (Excel).
"""
import os
import logging
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook

from config import (
    RAW_DIR, CLEANED_DIR, PRODUCT_FIELDS,
    get_excel_headers, get_field_keys, get_defaults,
)

logger = logging.getLogger(__name__)


def ensure_dirs():
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(CLEANED_DIR, exist_ok=True)


# ─── Cleaning ───────────────────────────────────────────────

def clean_record(record):
    cleaned = get_defaults()
    for key in get_field_keys():
        if key in record and record[key] is not None:
            cleaned[key] = record[key]
    for key, _, py_type, _ in PRODUCT_FIELDS:
        if py_type is str and cleaned[key] is not None:
            cleaned[key] = str(cleaned[key]).strip()
    try:
        cleaned["price"] = float(cleaned["price"]) if cleaned["price"] not in (None, "") else 0.0
    except (ValueError, TypeError):
        cleaned["price"] = 0.0
    try:
        cleaned["comment_count"] = int(float(str(cleaned["comment_count"]).replace(",", "")))
    except (ValueError, TypeError):
        cleaned["comment_count"] = 0
    try:
        cleaned["original_price"] = float(cleaned["original_price"]) if cleaned["original_price"] not in (None, "") else None
    except (ValueError, TypeError):
        cleaned["original_price"] = None
    try:
        cleaned["rating"] = float(cleaned["rating"]) if cleaned["rating"] not in (None, "") else None
    except (ValueError, TypeError):
        cleaned["rating"] = None
    cleaned["scraped_at"] = datetime.now(timezone.utc).isoformat()
    if not cleaned.get("name"):
        return None
    return cleaned


def clean_records(records):
    return [c for r in records if (c := clean_record(r)) is not None]


# ─── Merging ────────────────────────────────────────────────

def merge_data(existing, new_records, dedup_key="product_url"):
    lookup = {}
    for rec in existing:
        key_val = rec.get(dedup_key, "")
        if key_val:
            lookup[key_val] = rec
    for rec in new_records:
        key_val = rec.get(dedup_key, "")
        if key_val and key_val in lookup:
            lookup[key_val] = rec
        else:
            lookup[key_val or id(rec)] = rec
    return list(lookup.values())


# ─── Write ──────────────────────────────────────────────────

def _write_records_to_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.append(get_excel_headers())
    keys = get_field_keys()
    for rec in records:
        ws.append([rec.get(k, "") for k in keys])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    return wb


def save_raw(records, filename=None):
    ensure_dirs()
    if not filename:
        filename = f"raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(RAW_DIR, filename)
    wb = _write_records_to_workbook(records)
    wb.save(filepath)
    logger.info("Saved %d raw records to %s", len(records), filepath)
    return os.path.abspath(filepath)


def save_cleaned(records, filename="products.xlsx"):
    ensure_dirs()
    filepath = os.path.join(CLEANED_DIR, filename)
    wb = _write_records_to_workbook(records)
    wb.save(filepath)
    logger.info("Saved %d cleaned records to %s", len(records), filepath)
    return os.path.abspath(filepath)


# ─── Read ───────────────────────────────────────────────────

def read_excel(filepath):
    if not os.path.exists(filepath):
        return []
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    header_to_key = {header: key for key, header, _, _ in PRODUCT_FIELDS}
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 1:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        rec = get_defaults()
        for i, header in enumerate(headers):
            if header in header_to_key and i < len(row) and row[i] is not None:
                rec[header_to_key[header]] = row[i]
        if rec.get("name") and str(rec["name"]).strip():
            rec["created_at"] = rec.get("scraped_at", "")
            records.append(rec)
    logger.info("Read %d records from %s", len(records), filepath)
    return records


def read_cleaned(filename="products.xlsx"):
    filepath = os.path.join(CLEANED_DIR, filename)
    return read_excel(filepath)


# ─── Clear ──────────────────────────────────────────────────

def clear_products():
    filepath = os.path.join(CLEANED_DIR, "products.xlsx")
    if os.path.exists(filepath):
        os.remove(filepath)
        logger.info("Cleared products.xlsx")


# ─── Utilities ──────────────────────────────────────────────

def list_raw_files():
    if not os.path.exists(RAW_DIR):
        return []
    return sorted(f for f in os.listdir(RAW_DIR) if f.endswith(".xlsx"))


def get_stats(records):
    if not records:
        return {"total": 0, "price_range": {"min": None, "max": None, "avg": None},
                "platform_stats": [], "rating_dist": []}
    prices = [r["price"] for r in records if r.get("price") and r["price"] > 0]
    platforms = {}
    for r in records:
        p = r.get("platform") or "未知"
        platforms.setdefault(p, {"sum": 0, "cnt": 0})
        if r.get("price") and r["price"] > 0:
            platforms[p]["sum"] += r["price"]
            platforms[p]["cnt"] += 1
        else:
            platforms[p]["cnt"] += 1
    platform_stats = [
        {"platform": k, "count": v["cnt"], "avg_price": round(v["sum"] / v["cnt"], 2) if v["cnt"] else 0}
        for k, v in platforms.items()
    ]
    platform_stats.sort(key=lambda x: x["count"], reverse=True)
    return {
        "total": len(records),
        "price_range": {
            "min": round(min(prices), 2) if prices else None,
            "max": round(max(prices), 2) if prices else None,
            "avg": round(sum(prices) / len(prices), 2) if prices else None,
        },
        "platform_stats": platform_stats,
        "rating_dist": [],
    }
