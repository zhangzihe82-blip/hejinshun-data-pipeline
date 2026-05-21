"""
Excel → ECharts 代码生成器
独立应用，左侧上传Excel，右侧生成代码
"""
import os
import sys
import json
import logging
import webbrowser
import openpyxl
from flask import Flask, render_template, request, jsonify

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

if getattr(sys, 'frozen', False):
    _template_dir = os.path.join(sys._MEIPASS, 'templates')
    app = Flask(__name__, template_folder=_template_dir)
else:
    app = Flask(__name__)

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# 文件名关键词 → 图表类型映射
CHART_KEYWORDS = {
    '饼图': 'pie',
    '漏斗图': 'funnel',
    '柱状图': 'bar',
    '气泡图': 'scatter',
    '词云': 'wordCloud',
    '雷达图': 'radar',
    '矩形树图': 'treemap',
    '环形图': 'donut',
}

# ─── 页面 ─────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('generator.html')


# ─── 处理上传的 Excel 文件 ─────────────────────────────────

@app.route('/api/process', methods=['POST'])
def process_files():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '未收到文件'}), 400

    charts = []
    for f in files:
        if not f.filename.endswith(('.xlsx', '.xls')):
            continue
        try:
            chart_config = _parse_excel(f)
            if chart_config:
                charts.append(chart_config)
        except Exception as e:
            logger.error(f'解析 {f.filename} 失败: {e}')
            charts.append({
                'id': f'error_{len(charts)}',
                'type': 'bar',
                'title': f'{f.filename} (解析失败)',
                'error': str(e),
                'data': []
            })

    config = {
        'title': '数据分析大屏',
        'charts': charts
    }

    code = json.dumps(config, ensure_ascii=False, indent=2)
    return jsonify({'success': True, 'config': config, 'code': code, 'count': len(charts)})


def _parse_excel(file_storage):
    """解析单个 Excel 文件，生成 ECharts 图表配置"""
    filename = file_storage.filename
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = wb.active

    # 识别图表类型
    chart_type = 'bar'  # 默认柱状图
    for kw, ct in CHART_KEYWORDS.items():
        if kw in filename:
            chart_type = ct
            break

    # 提取图表标题（去掉后缀和扩展名）
    title = filename.replace('.xlsx', '').replace('.xls', '')
    # 去掉括号中的图表类型提示
    import re
    title = re.sub(r'[（(][^)）]*[)）]', '', title).strip()

    # 读取数据
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = [v for v in row if v is not None]
        vals = [v for v in vals if str(v).strip()]
        if vals:
            rows.append(vals)

    if len(rows) < 2:
        return None

    headers = [str(r) for r in rows[0]]
    data_rows = rows[1:]

    # 根据图表类型生成配置
    chart_id = f'chart_{filename.replace(".xlsx","").replace(".xls","").replace(" ","_")}'

    if chart_type in ('pie', 'donut'):
        return _make_pie_donut(chart_id, chart_type, title, headers, data_rows)
    elif chart_type == 'funnel':
        return _make_funnel(chart_id, title, headers, data_rows)
    elif chart_type == 'bar':
        return _make_bar(chart_id, title, headers, data_rows)
    elif chart_type == 'scatter':
        return _make_scatter(chart_id, title, headers, data_rows)
    elif chart_type == 'wordCloud':
        return _make_wordcloud(chart_id, title, headers, data_rows)
    elif chart_type == 'radar':
        return _make_radar(chart_id, title, headers, data_rows)
    elif chart_type == 'treemap':
        return _make_treemap(chart_id, title, headers, data_rows)

    return None


def _clean_num(v):
    """清洗数值"""
    try:
        return float(str(v).replace('%', '').replace(',', '').replace('¥', '').strip())
    except:
        return 0


def _make_pie_donut(chart_id, chart_type, title, headers, data_rows):
    data = []
    for row in data_rows:
        if len(row) >= 2:
            data.append({'name': str(row[0]), 'value': _clean_num(row[1])})
            if len(row) >= 3 and row[2]:
                data[-1]['extra'] = str(row[2])
    return {
        'id': chart_id, 'type': chart_type, 'title': title,
        'seriesName': headers[1] if len(headers) > 1 else '占比',
        'data': data
    }


def _make_funnel(chart_id, title, headers, data_rows):
    data = []
    for row in data_rows:
        if len(row) >= 2:
            data.append({'name': str(row[0]), 'value': _clean_num(row[1])})
    # 漏斗图数据按值降序
    data.sort(key=lambda x: x['value'], reverse=True)
    return {
        'id': chart_id, 'type': 'funnel', 'title': title,
        'seriesName': headers[1] if len(headers) > 1 else '占比',
        'data': data
    }


def _make_bar(chart_id, title, headers, data_rows):
    data = []
    for row in data_rows:
        if len(row) >= 2:
            name = str(row[0])
            val = _clean_num(row[1])
            if name and val is not None:
                data.append({'name': name, 'value': val})
    return {
        'id': chart_id, 'type': 'bar', 'title': title,
        'xAxis': headers[0] if headers else '类别',
        'yAxis': headers[1] if len(headers) > 1 else '数值',
        'data': data
    }


def _make_scatter(chart_id, title, headers, data_rows):
    # 气泡图: 列1=名称, 列2=X, 列3=Y, 列4=气泡大小
    data = []
    for row in data_rows:
        if len(row) >= 3:
            name = str(row[0])
            x = _clean_num(row[1])
            y = _clean_num(row[2])
            size = _clean_num(row[3]) if len(row) >= 4 else 10
            if name:
                data.append({'name': name, 'value': [x, y, size]})
    return {
        'id': chart_id, 'type': 'scatter', 'title': title,
        'xAxis': headers[1] if len(headers) > 1 else 'X',
        'yAxis': headers[2] if len(headers) > 2 else 'Y',
        'data': data
    }


def _make_wordcloud(chart_id, title, headers, data_rows):
    data = []
    for row in data_rows:
        if len(row) >= 2:
            name = str(row[0])
            val = _clean_num(row[1])
            if name:
                # 词云权重按比例缩放到合适范围
                data.append({'name': name, 'value': val})
    return {
        'id': chart_id, 'type': 'wordCloud', 'title': title,
        'data': data
    }


def _make_radar(chart_id, title, headers, data_rows):
    # 雷达图: 列1=维度名, 列2=数值
    indicators = []
    values = []
    for row in data_rows:
        if len(row) >= 2:
            name = str(row[0])
            val = _clean_num(row[1])
            if name:
                indicators.append({'name': name, 'max': 5})  # 固定5分制
                values.append(val)
    return {
        'id': chart_id, 'type': 'radar', 'title': title,
        'seriesName': title,
        'indicators': indicators,
        'data': [{'name': title, 'value': values}]
    }


def _make_treemap(chart_id, title, headers, data_rows):
    data = []
    for row in data_rows:
        if len(row) >= 2:
            data.append({'name': str(row[0]), 'value': _clean_num(row[1])})
    return {
        'id': chart_id, 'type': 'treemap', 'title': title,
        'data': data
    }


# ─── 启动 ─────────────────────────────────────────────────

if __name__ == '__main__':
    port = 5002
    url = f'http://127.0.0.1:{port}'
    logger.info(f'代码生成器启动: {url}')
    webbrowser.open(url)
    app.run(debug=False, host='127.0.0.1', port=port, threaded=True)
