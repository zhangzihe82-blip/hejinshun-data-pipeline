"""
Excel输出器
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from typing import Dict, Any, List


class ExcelWriter:
    """Excel输出器

    生成兼容和金顺平台的Excel文件
    """

    # 和金顺平台兼容的列标题映射
    HEADER_MAP = {
        'name': 'Product Name',
        'price': 'Price',
        'original_price': 'Original Price',
        'platform': 'Platform',
        'rating': 'Rating',
        'comment_count': 'Comment Count',
        'image_url': 'Image URL',
        'product_url': 'Product URL',
        'category': 'Category',
        'scraped_at': 'Scraped At',
        'created_at': 'Created At',
        # 新增用户画像字段
        'user_age': 'User Age',
        'user_gender': 'User Gender',
        'user_region': 'User Region'
    }

    # 列顺序
    COLUMN_ORDER = [
        'name', 'price', 'original_price', 'platform',
        'rating', 'comment_count', 'category',
        'user_age', 'user_gender', 'user_region',
        'image_url', 'product_url', 'scraped_at', 'created_at'
    ]

    def __init__(self):
        """初始化输出器"""
        self.header_style = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write(self, data: List[Dict[str, Any]], output_path: Path, config: Dict[str, Any] = None):
        """写入Excel文件

        Args:
            data: 数据列表
            output_path: 输出路径
            config: 配置字典（可选）
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Products'

        # 写入标题行
        headers = [self.HEADER_MAP.get(col, col) for col in self.COLUMN_ORDER]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_style
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 写入数据行
        for row_idx, record in enumerate(data, 2):
            for col_idx, field_name in enumerate(self.COLUMN_ORDER, 1):
                value = record.get(field_name)

                # 格式化
                if field_name == 'price' and value is not None:
                    value = round(float(value), 2)
                elif field_name == 'original_price' and value is not None:
                    value = round(float(value), 2)
                elif field_name == 'rating' and value is not None:
                    value = round(float(value), 1)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border

        # 调整列宽
        column_widths = {
            'A': 40,  # Product Name
            'B': 12,  # Price
            'C': 12,  # Original Price
            'D': 15,  # Platform
            'E': 8,   # Rating
            'F': 12,  # Comment Count
            'G': 12,  # Category
            'H': 10,  # User Age
            'I': 12,  # User Gender
            'J': 10,  # User Region
            'K': 40,  # Image URL
            'L': 35,  # Product URL
            'M': 20   # Scraped At
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 保存文件
        wb.save(output_path)
