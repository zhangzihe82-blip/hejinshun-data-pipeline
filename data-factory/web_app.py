"""
数据制造工厂 - Web界面
独立运行的Web应用
"""
import os
import sys
import json
import yaml
import threading
import logging
from datetime import datetime
from pathlib import Path

# 添加当前目录到Python路径
_MODULE_DIR = Path(__file__).parent
sys.path.insert(0, str(_MODULE_DIR))

# 数据目录基准: 打包后指向 exe 所在目录，开发时指向项目根目录
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = _MODULE_DIR.parent

from flask import Flask, render_template, request, jsonify, send_file

logger = logging.getLogger(__name__)

# 创建Flask应用 (模板/静态文件仍从模块目录加载)
if getattr(sys, 'frozen', False):
    _template_dir = Path(sys._MEIPASS) / 'data-factory' / 'templates'
    _static_dir = Path(sys._MEIPASS) / 'data-factory' / 'static'
else:
    _template_dir = _MODULE_DIR / 'templates'
    _static_dir = _MODULE_DIR / 'static'

app = Flask(__name__,
            template_folder=str(_template_dir),
            static_folder=str(_static_dir))

# 全局状态
_generate_status = {
    'running': False,
    'total': 0,
    'current': 0,
    'message': '就绪',
    'output_file': None
}
_status_lock = threading.Lock()


def reset_status():
    """重置生成状态"""
    with _status_lock:
        _generate_status['running'] = False
        _generate_status['total'] = 0
        _generate_status['current'] = 0
        _generate_status['message'] = '就绪'
        _generate_status['output_file'] = None


def load_config(config_name='ecommerce'):
    """加载配置文件"""
    config_path = _MODULE_DIR / 'config' / f'{config_name}.yaml'
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    return {}


def get_field_defaults():
    """获取字段默认值"""
    return {
        'name': {'type': 'string', 'method': 'template', 'template': '{brand} {category} {feature}'},
        'price': {'type': 'float', 'method': 'lognormal', 'mean': 6.0, 'sigma': 1.0, 'min': 1.0, 'max': 50000.0},
        'original_price': {'type': 'float', 'method': 'derived', 'null_probability': 0.15},
        'platform': {'type': 'string', 'method': 'choice', 'choices': ['京东', '京东自营', '什么值得买'], 'weights': [0.35, 0.45, 0.20]},
        'rating': {'type': 'float', 'method': 'truncated_normal', 'mean': 4.2, 'sigma': 0.5, 'min': 1.0, 'max': 5.0},
        'comment_count': {'type': 'integer', 'method': 'poisson', 'lambda': 800, 'min': 0},
        'category': {'type': 'string', 'method': 'choice', 'choices': ['手机', '电脑', '家电', '服装', '食品', '图书', '美妆', '家居']},
        'image_url': {'type': 'string', 'method': 'template', 'template': 'https://img14.360buyimg.com/n1/{id}.jpg'},
        'product_url': {'type': 'string', 'method': 'template', 'template': 'https://item.jd.com/{id}.html'},
        'scraped_at': {'type': 'datetime', 'method': 'uniform_range', 'start': '2024-01-01', 'end': '2024-12-31'}
    }


# ─── 页面路由 ─────────────────────────────────────────────

@app.route('/')
def index():
    """主页面"""
    return render_template('factory.html')


# ─── API路由 ──────────────────────────────────────────────

@app.route('/api/defaults')
def api_defaults():
    """获取默认配置"""
    defaults = get_field_defaults()
    return jsonify(defaults)


@app.route('/api/config/<name>')
def api_config(name):
    """获取指定配置"""
    import re
    # 防止路径遍历攻击：只允许安全的文件名字符
    if not re.match(r'^[a-zA-Z0-9_-]+$', name):
        return jsonify({'error': '无效的配置名称'}), 400
    config = load_config(name)
    return jsonify(config)


@app.route('/api/parse-conclusion', methods=['POST'])
def api_parse_conclusion():
    """解析结论语句"""
    data = request.get_json() or {}
    conclusion_text = data.get('conclusion', '').strip()

    if not conclusion_text:
        return jsonify({'error': '请输入结论语句'}), 400

    try:
        from conclusion_engine import ConclusionDrivenGenerator
        generator = ConclusionDrivenGenerator()
        conclusion = generator.parser.parse(conclusion_text)

        return jsonify({
            'success': True,
            'parsed': {
                'category': conclusion.category,
                'platform': conclusion.platform,
                'metric': conclusion.metric,
                'direction': conclusion.direction.value if conclusion.direction else None,
                'comparison': conclusion.comparison,
                'confidence': conclusion.confidence
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-from-conclusion', methods=['POST'])
def api_generate_from_conclusion():
    """根据结论生成数据"""
    global _generate_status

    with _status_lock:
        if _generate_status['running']:
            return jsonify({'error': '已有生成任务在运行中'}), 400
        _generate_status['running'] = True
        _generate_status['total'] = 0
        _generate_status['current'] = 0
        _generate_status['message'] = '解析结论...'

    data = request.get_json() or {}
    conclusion_text = data.get('conclusion', '').strip()
    count = max(1, min(int(data.get('count', 100)), 10000))
    seed = data.get('seed')
    output_format = data.get('format', 'excel')

    def _task():
        global _generate_status
        try:
            from conclusion_engine import generate_from_conclusion
            from generators import DataGenerator
            from output import ExcelWriter, JSONWriter, CSVWriter

            # 根据结论生成配置
            config, explanation = generate_from_conclusion(conclusion_text, count, seed)

            with _status_lock:
                _generate_status['message'] = '正在生成数据...'
                _generate_status['total'] = count

            # 创建生成器
            generator = DataGenerator(config, seed=seed)

            # 生成数据
            records = generator.generate(count)

            with _status_lock:
                _generate_status['current'] = len(records)
                _generate_status['message'] = '正在保存文件...'

            # 保存文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = BASE_DIR / 'output'
            output_dir.mkdir(exist_ok=True)

            if output_format == 'json':
                output_file = output_dir / f'data_{timestamp}.json'
                writer = JSONWriter()
                writer.write(records, output_file)
            elif output_format == 'csv':
                output_file = output_dir / f'data_{timestamp}.csv'
                writer = CSVWriter()
                writer.write(records, output_file)
            else:
                output_file = output_dir / f'data_{timestamp}.xlsx'
                writer = ExcelWriter()
                writer.write(records, output_file, config)

            with _status_lock:
                _generate_status['running'] = False
                _generate_status['message'] = f'完成！共生成 {len(records)} 条数据'
                _generate_status['output_file'] = str(output_file)

        except Exception as e:
            logger.exception('生成数据出错')
            with _status_lock:
                _generate_status['running'] = False
                _generate_status['message'] = f'出错: {str(e)[:50]}'

    t = threading.Thread(target=_task, daemon=True)
    t.start()
    return jsonify({'success': True, 'message': '生成任务已启动', 'parsed': None})


@app.route('/api/generate-structured', methods=['POST'])
def api_generate_structured():
    """根据结构化参数生成数据"""
    global _generate_status

    with _status_lock:
        if _generate_status['running']:
            return jsonify({'error': '已有生成任务在运行中'}), 400
        _generate_status['running'] = True
        _generate_status['total'] = 0
        _generate_status['current'] = 0
        _generate_status['message'] = '构建配置...'

    data = request.get_json() or {}
    count = max(1, min(int(data.get('count', 500)), 10000))
    seed = data.get('seed')
    output_format = data.get('format', 'excel')

    def _task():
        global _generate_status
        try:
            from conclusion_engine import ConclusionParser, DataRuleGenerator, ConclusionDrivenGenerator, TrendDirection
            from generators import DataGenerator
            from output import ExcelWriter, JSONWriter, CSVWriter

            # 使用结构化参数解析
            parser = ConclusionParser()
            conclusion = parser.parse_structured(data)

            # 生成规则
            rule_generator = DataRuleGenerator()
            rules = rule_generator.generate_rules(conclusion)

            # 构建完整配置
            generator = ConclusionDrivenGenerator()
            config = generator.build_config(rules, conclusion, count, seed)

            with _status_lock:
                _generate_status['message'] = '正在生成数据...'
                _generate_status['total'] = count

            # 创建生成器并生成数据
            data_gen = DataGenerator(config, seed=seed)
            records = data_gen.generate(count)

            with _status_lock:
                _generate_status['current'] = len(records)
                _generate_status['message'] = '正在保存文件...'

            # 保存文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = BASE_DIR / 'output'
            output_dir.mkdir(exist_ok=True)

            if output_format == 'json':
                output_file = output_dir / f'data_{timestamp}.json'
                writer = JSONWriter()
                writer.write(records, output_file)
            elif output_format == 'csv':
                output_file = output_dir / f'data_{timestamp}.csv'
                writer = CSVWriter()
                writer.write(records, output_file)
            else:
                output_file = output_dir / f'data_{timestamp}.xlsx'
                writer = ExcelWriter()
                writer.write(records, output_file, config)

            with _status_lock:
                _generate_status['running'] = False
                _generate_status['message'] = f'完成！共生成 {len(records)} 条数据'
                _generate_status['output_file'] = str(output_file)

        except Exception as e:
            logger.exception('生成数据出错')
            import traceback
            error_detail = traceback.format_exc()
            logger.error(f'详细错误: {error_detail}')
            with _status_lock:
                _generate_status['running'] = False
                _generate_status['message'] = f'出错: {str(e)[:100]}'

    t = threading.Thread(target=_task, daemon=True)
    t.start()
    return jsonify({'success': True, 'message': '生成任务已启动'})


@app.route('/api/generate', methods=['POST'])
def api_generate():
    """生成数据"""
    global _generate_status

    with _status_lock:
        if _generate_status['running']:
            return jsonify({'error': '已有生成任务在运行中'}), 400
        _generate_status['running'] = True
        _generate_status['total'] = 0
        _generate_status['current'] = 0
        _generate_status['message'] = '准备生成...'

    data = request.get_json() or {}
    count = max(1, min(int(data.get('count', 100)), 10000))
    seed = data.get('seed')
    field_config = data.get('fields', {})
    output_format = data.get('format', 'excel')

    def _task():
        global _generate_status
        try:
            from generators import DataGenerator
            from output import ExcelWriter, JSONWriter, CSVWriter

            # 合并默认配置
            defaults = get_field_defaults()
            merged_config = {'fields': {}}
            for field_name, default_cfg in defaults.items():
                if field_name in field_config:
                    merged_config['fields'][field_name] = {**default_cfg, **field_config[field_name]}
                else:
                    merged_config['fields'][field_name] = default_cfg

            # 添加约束和关联
            merged_config['constraints'] = [
                {'rule': 'original_price >= price'},
                {'rule': '1.0 <= rating <= 5.0'},
                {'rule': 'comment_count >= 0'}
            ]
            merged_config['correlations'] = [
                {'fields': ['rating', 'comment_count'], 'coefficient': 0.35}
            ]

            with _status_lock:
                _generate_status['message'] = '正在生成数据...'
                _generate_status['total'] = count

            # 创建生成器
            generator = DataGenerator(merged_config, seed=seed)

            # 生成数据
            records = generator.generate(count)

            with _status_lock:
                _generate_status['current'] = len(records)
                _generate_status['message'] = '正在保存文件...'

            # 保存文件
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_dir = BASE_DIR / 'output'
            output_dir.mkdir(exist_ok=True)

            if output_format == 'json':
                output_file = output_dir / f'data_{timestamp}.json'
                writer = JSONWriter()
                writer.write(records, output_file)
            elif output_format == 'csv':
                output_file = output_dir / f'data_{timestamp}.csv'
                writer = CSVWriter()
                writer.write(records, output_file)
            else:
                output_file = output_dir / f'data_{timestamp}.xlsx'
                writer = ExcelWriter()
                writer.write(records, output_file, merged_config)

            with _status_lock:
                _generate_status['running'] = False
                _generate_status['message'] = f'完成！共生成 {len(records)} 条数据'
                _generate_status['output_file'] = str(output_file)

        except Exception as e:
            logger.exception('生成数据出错')
            with _status_lock:
                _generate_status['running'] = False
                _generate_status['message'] = f'出错: {str(e)[:50]}'

    t = threading.Thread(target=_task, daemon=True)
    t.start()
    return jsonify({'success': True, 'message': '生成任务已启动'})


@app.route('/api/status')
def api_status():
    """获取生成状态"""
    with _status_lock:
        return jsonify(dict(_generate_status))


@app.route('/api/reset', methods=['POST'])
def api_reset():
    """重置生成状态"""
    reset_status()
    return jsonify({'success': True, 'message': '状态已重置'})


@app.route('/api/download')
def api_download():
    """下载生成的文件"""
    with _status_lock:
        output_file = _generate_status.get('output_file')

    if not output_file or not Path(output_file).exists():
        return jsonify({'error': '文件不存在'}), 404

    return send_file(output_file, as_attachment=True)


@app.route('/api/export', methods=['POST'])
def api_export():
    """导出到和金顺平台"""
    with _status_lock:
        output_file = _generate_status.get('output_file')

    if not output_file or not Path(output_file).exists():
        return jsonify({'error': '请先生成数据'}), 400

    try:
        import openpyxl
        from datetime import datetime

        # 读取生成的数据
        wb = openpyxl.load_workbook(output_file, data_only=True)
        ws = wb.active

        # 获取标题和数据
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        if len(rows) < 2:
            return jsonify({'error': '生成的数据为空'}), 400

        headers = list(rows[0])
        data_rows = rows[1:]

        # 和金顺平台数据文件路径 (BASE_DIR 已指向项目根目录/exe所在目录)
        hejinshun_path = BASE_DIR / 'data' / 'cleaned' / 'products.xlsx'
        hejinshun_path.parent.mkdir(parents=True, exist_ok=True)

        # 如果目标文件存在，追加数据
        if hejinshun_path.exists():
            # 不使用data_only，保留原有数据格式
            target_wb = openpyxl.load_workbook(hejinshun_path)
            target_ws = target_wb.active

            # 获取当前行数
            original_count = target_ws.max_row - 1

            # 追加数据（跳过标题行）
            for row in data_rows:
                target_ws.append(row)

            target_wb.save(hejinshun_path)
            total_count = target_ws.max_row - 1

            logger.info(f"导出完成: 新增 {len(data_rows)} 条，原有 {original_count} 条，总计 {total_count} 条")
        else:
            # 直接复制文件
            import shutil
            shutil.copy(output_file, hejinshun_path)
            total_count = len(data_rows)
            logger.info(f"创建新文件: 导出 {total_count} 条数据")

        return jsonify({
            'success': True,
            'message': f'已导出 {len(data_rows)} 条数据到和金顺平台，当前共 {total_count} 条数据',
            'path': str(hejinshun_path),
            'new_count': len(data_rows),
            'total_count': total_count
        })
    except Exception as e:
        logger.exception('导出数据出错')
        return jsonify({'error': f'导出失败: {str(e)}'}), 500


@app.route('/shutdown', methods=['POST'])
def shutdown():
    """关闭Flask服务器"""
    func = request.environ.get('werkzeug.server.shutdown')
    if func is not None:
        func()
    return jsonify({'success': True, 'message': '服务正在关闭'})


def run_app(port=5007, open_browser=True, stop_event=None):
    """运行应用"""
    import webbrowser
    import threading

    url = f'http://127.0.0.1:{port}'
    logger.info(f'启动数据制造工厂: {url}')

    if open_browser:
        webbrowser.open(url)

    # 如果有stop_event，启动一个监控线程，当事件触发时请求shutdown
    if stop_event is not None:
        def _watch_stop():
            stop_event.wait()
            try:
                import urllib.request
                urllib.request.urlopen(f'http://127.0.0.1:{port}/shutdown', timeout=2)
            except Exception:
                pass
        watcher = threading.Thread(target=_watch_stop, daemon=True)
        watcher.start()

    app.run(debug=False, host='127.0.0.1', port=port, threaded=True)


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    run_app()
