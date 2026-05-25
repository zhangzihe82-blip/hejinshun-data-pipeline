"""
数据制造工厂 - CLI入口
生成具有统计真实性的电商数据
"""
import click
import yaml
from pathlib import Path
from datetime import datetime
from generators import DataGenerator
from validators import DataValidator
from output import ExcelWriter, JSONWriter, CSVWriter


@click.group()
def cli():
    """数据制造工厂 - 生成统计真实的数据"""
    pass


@cli.command()
@click.option('--count', '-n', default=100, help='生成记录数量')
@click.option('--config', '-c', 'config_path', default='config/ecommerce.yaml', help='配置文件路径')
@click.option('--output', '-o', 'output_path', required=True, help='输出文件路径')
@click.option('--seed', '-s', 'random_seed', default=None, type=int, help='随机种子(可重现)')
@click.option('--format', '-f', 'output_format', type=click.Choice(['excel', 'json', 'csv']), default='excel', help='输出格式')
def generate(count, config_path, output_path, random_seed, output_format):
    """生成数据"""
    # 加载配置
    config_file = Path(config_path)
    if not config_file.exists():
        click.echo(f"错误: 配置文件不存在 - {config_path}", err=True)
        return 1

    with open(config_file, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    click.echo(f"正在生成 {count} 条数据...")

    # 创建生成器
    generator = DataGenerator(config, seed=random_seed)

    # 生成数据
    data = generator.generate(count)

    click.echo(f"数据生成完成，共 {len(data)} 条记录")

    # 输出
    output_file = Path(output_path)
    if output_format == 'excel' or output_file.suffix in ['.xlsx', '.xls']:
        writer = ExcelWriter()
        writer.write(data, output_file, config)
    elif output_format == 'json' or output_file.suffix == '.json':
        writer = JSONWriter()
        writer.write(data, output_file)
    else:
        writer = CSVWriter()
        writer.write(data, output_file)

    click.echo(f"数据已保存到: {output_file}")


@cli.command()
@click.option('--input', '-i', 'input_path', required=True, help='输入文件路径')
@click.option('--config', '-c', 'config_path', default='config/ecommerce.yaml', help='配置文件路径')
def validate(input_path, config_path):
    """验证数据质量"""
    # 加载配置
    config_file = Path(config_path)
    if not config_file.exists():
        click.echo(f"错误: 配置文件不存在 - {config_path}", err=True)
        return 1

    with open(config_file, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)

    # 读取数据
    input_file = Path(input_path)
    if input_file.suffix in ['.xlsx', '.xls']:
        import openpyxl
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            data.append(record)
    else:
        click.echo("仅支持Excel格式验证", err=True)
        return 1

    click.echo(f"正在验证 {len(data)} 条记录...")

    # 验证
    validator = DataValidator(config)
    report = validator.validate(data)

    # 输出报告
    click.echo("\n=== 验证报告 ===")
    click.echo(f"总记录数: {report['total_records']}")
    click.echo(f"一致性通过率: {report['consistency_rate']:.1%}")
    click.echo(f"唯一性通过率: {report['uniqueness_rate']:.1%}")

    if report['issues']:
        click.echo("\n发现的问题:")
        for issue in report['issues'][:10]:
            click.echo(f"  - {issue}")
        if len(report['issues']) > 10:
            click.echo(f"  ... 还有 {len(report['issues']) - 10} 个问题")

    return 0 if report['is_valid'] else 1


@cli.command()
def info():
    """显示配置信息"""
    click.echo("数据制造工厂 v1.0.0")
    click.echo("支持的字段: name, price, original_price, platform, rating, comment_count, image_url, product_url, category, scraped_at")


if __name__ == '__main__':
    cli()
