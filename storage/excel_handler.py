"""
Excel处理子模块
负责Excel文件的读写操作
"""
import os
import logging
from datetime import datetime

from openpyxl import Workbook, load_workbook

from config import (
    RAW_DIR, CLEANED_DIR, PRODUCT_FIELDS,
    get_excel_headers, get_field_keys,
)

logger = logging.getLogger(__name__)


def ensure_dirs():
    """确保数据目录存在"""
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(CLEANED_DIR, exist_ok=True)


# ─── 写入 ─────────────────────────────────────────────────────

def _write_records_to_workbook(records):
    """将记录写入工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.append(get_excel_headers())
    keys = get_field_keys()
    for rec in records:
        ws.append([rec.get(k, "") for k in keys])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    return wb


def save_raw(records, filename=None):
    """保存原始数据"""
    ensure_dirs()
    if not filename:
        filename = f"raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(RAW_DIR, filename)
    wb = _write_records_to_workbook(records)
    wb.save(filepath)
    logger.info("Saved %d raw records to %s", len(records), filepath)
    return os.path.abspath(filepath)


def save_cleaned(records, filename="products.xlsx"):
    """保存清洗后的数据"""
    ensure_dirs()
    filepath = os.path.join(CLEANED_DIR, filename)
    wb = _write_records_to_workbook(records)
    wb.save(filepath)
    logger.info("Saved %d cleaned records to %s", len(records), filepath)
    return os.path.abspath(filepath)


# ─── 读取 ─────────────────────────────────────────────────────

def read_excel(filepath):
    """从Excel读取数据"""
    if not os.path.exists(filepath):
        return []
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    header_to_key = {header: key for key, header, _, _ in PRODUCT_FIELDS}
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 1:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    from config import get_defaults
    records = []
    for row in rows[1:]:
        rec = get_defaults()
        for i, header in enumerate(headers):
            if header in header_to_key and i < len(row) and row[i] is not None:
                rec[header_to_key[header]] = row[i]
        if rec.get("name") and str(rec["name"]).strip():
            rec["created_at"] = rec.get("scraped_at", "")
            records.append(rec)
    logger.info("Read %d records from %s", len(records), filepath)
    return records


def read_cleaned(filename="products.xlsx"):
    """读取清洗后的数据"""
    filepath = os.path.join(CLEANED_DIR, filename)
    return read_excel(filepath)


# ─── 清理 ─────────────────────────────────────────────────────

def clear_products():
    """清空产品数据"""
    filepath = os.path.join(CLEANED_DIR, "products.xlsx")
    if os.path.exists(filepath):
        os.remove(filepath)
        logger.info("Cleared products.xlsx")


# ─── 工具 ─────────────────────────────────────────────────────

def list_raw_files():
    """列出原始数据文件"""
    if not os.path.exists(RAW_DIR):
        return []
    return sorted(f for f in os.listdir(RAW_DIR) if f.endswith(".xlsx"))
